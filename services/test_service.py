import io
import json

from openpyxl.workbook import Workbook
from sqlalchemy import insert

from db.models import DAMAQuestion, DAMACase, TestResults, TestAnswer, Analytics
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from services.logger import logger
from services.redis_service import RedisService
from datetime import datetime
from sqlalchemy.future import select
from db.database import get_async_session
from db.models import DMARoles, DAMACompetency

async def prepare_test_data(selected_role: str, selected_comp: str):
    async with get_async_session() as session:
        try:
            theory_query = select(DAMAQuestion).where(
                DAMAQuestion.dama_role_name == selected_role,
                DAMAQuestion.dama_competence_name == selected_comp,
                DAMAQuestion.question_type == "Теория"
            ).distinct()
            theory_result = await session.execute(theory_query)
            theory_questions = theory_result.scalars().all()

            practice_query = select(DAMAQuestion).where(
                DAMAQuestion.dama_role_name == selected_role,
                DAMAQuestion.dama_competence_name == selected_comp,
                DAMAQuestion.question_type == "Практика"
            ).distinct()
            practice_result = await session.execute(practice_query)
            practice_questions = practice_result.scalars().all()


            case_query = select(DAMACase).where(
                DAMACase.dama_role_name == selected_role,
                DAMACase.dama_competence_name == selected_comp
            )
            case_result = await session.execute(case_query)
            case = case_result.scalars().first()

            selected_questions = balance_questions(theory_questions, practice_questions)

            return {
                'questions': selected_questions,
                'case': case,
                'total_questions': len(selected_questions),
                'has_case': bool(case)
            }
        except Exception as e:
            raise e

def balance_questions(theory_questions, practice_questions):
    if len(theory_questions) >= 5 and len(practice_questions) >= 5:
        selected_questions = theory_questions[:5] + practice_questions[:5]
    else:
        min_count = min(len(theory_questions), len(practice_questions))
        selected_questions = theory_questions[:min_count] + practice_questions[:min_count]
        remaining = 10 - len(selected_questions)
        if remaining > 0:
            if len(theory_questions) > min_count:
                selected_questions += theory_questions[min_count:min_count + remaining]
            else:
                selected_questions += practice_questions[min_count:min_count + remaining]
    return selected_questions


async def generate_test_report(user_id: int):
    redis_service = RedisService()
    answers = await redis_service.get_user_answers(user_id)
    metadata = await redis_service.get_user_metadata(user_id)
    analytics = await redis_service.load_analytics(user_id)

    total_score = 0.0
    valid_answers = 0
    total_prompt_tokens = 0
    total_completion_tokens = 0

    filtered_answers = []
    
    for answer in answers:
        if not answer.get('user_answer'):
            logger.warning(f"Skipping answer with empty text: {answer}")
            continue

        filtered_answers.append(answer)

    for answer in filtered_answers:
        try:
            score = float(answer.get('score', 0)) if isinstance(answer, dict) else 0
            total_score += score
            valid_answers += 1
        except (ValueError, TypeError):
            continue

    for analytic in analytics:
        total_prompt_tokens += analytic.get('prompt_tokens', 0)
        total_completion_tokens += analytic.get('completion_tokens', 0)

    avg = round(total_score / valid_answers, 2) if valid_answers > 0 else 0.0
    is_expert = avg >= 4.5
    model = await redis_service.load_selected_ai_model()

    async with get_async_session() as session:
        try:

            test_result = {
                'user_id': user_id,
                'dama_role': metadata.get('selected_role', ''),
                'dama_competence': metadata.get('selected_comp', ''),
                'total_score': avg,
                'is_expert': is_expert,
                'test_date': datetime.utcnow()
            }
            result = await session.execute(
                insert(TestResults).values(**test_result).returning(TestResults.id)
            )
            test_result_id = result.scalar_one()

            for answer in filtered_answers:
                logger.debug(f"Answer: {answer}")

                answer_data = {
                    'test_result_id': test_result_id,
                    'question_id': answer.get('question_id'),
                    'case_id': answer.get('case_id'),
                    'answer_text': answer.get('user_answer', ''),
                    'score': answer.get('score', 0),
                    'feedback': json.dumps(answer.get('feedback', {}))
                }
                await session.execute(insert(TestAnswer).values(**answer_data))

            if model:
                analytics_data = {
                    'test_result_id': test_result_id,
                    'prompt_tokens': total_prompt_tokens,
                    'completion_tokens': total_completion_tokens,
                    'total_tokens': total_prompt_tokens + total_completion_tokens,
                    'model': model
                }
                await session.execute(insert(Analytics).values(**analytics_data))

            await session.commit()
        except Exception as e:
            await session.rollback()
            logger.error(f"Error saving test results to DB: {e}")
            raise

    wb = Workbook()
    ws = wb.active
    
    if ws is not None: 
        ws.title = "DAMA Assessment Report"
    else:
        ws = wb.create_sheet("DAMA Assessment Report")

    header_font = Font(bold=True, name='Century Gothic', size=12)
    body_font = Font(name='Century Gothic', size=11)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color='D5FD7B', end_color='D5FD7B', fill_type='solid')
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))


    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = "Результаты оценки уровня владения компетенциями по управлению данными в соответствии с фреймворком DAMА"
    title_cell.font = header_font
    title_cell.alignment = center_alignment
    title_cell.fill = header_fill
    title_cell.border = thin_border

    ws.append([None])

    meta_rows = [
        ["Дата и время тестирования", "", "", "", ""],
        ["ФИО тестируемого", "", "", "", ""],
        ["Роль", "", "", "", ""],
        ["Компетенция", "", "", "", ""],
        ["Средняя оценка по компетенции", "", "", "", ""],
        ["Экспертность в управлении данными DAMA (порог экспертности ≥ 4.5)", "", "", "", ""],
        [None, None, None, None, None],
        ["Область знаний/Основные работы", "Вопрос", "Пользовательский ответ", "Рекомендуемые материалы для изучения", "Оценка (1-5)"]
    ]

    for row_idx, row in enumerate(meta_rows, start=3):
        for col_idx, value in enumerate(row, start=1):
            if value is not None:
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = body_font
                cell.alignment = center_alignment
                if row_idx == 10:
                    cell.font = Font(bold=True, name='Century Gothic', size=11)
                    cell.fill = header_fill
                if value != "":
                    cell.border = thin_border

    ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['B4'] = metadata.get('user_name', '')
    ws['B5'] = metadata.get('selected_role', '')
    ws['B6'] = metadata.get('selected_comp', '')
    ws['B7'] = avg
    ws['B8'] = "Да" if is_expert else "Нет"

    for row in range(3, 9):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell.font = body_font
                cell.alignment = center_alignment
                cell.border = thin_border

    for answer in filtered_answers:
        feedback = answer.get('feedback', {})
        recommendations = feedback.get('recommendations', '')
        if isinstance(recommendations, list):
            recommendations = " ".join(recommendations)

        row = [
            answer.get('knowledge_area', ''),
            answer.get('question', ''),
            answer.get('user_answer', ''),
            recommendations,
            float(answer.get('score', 0))
        ]
        ws.append(row)


        for col in range(1, 6):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = body_font
            cell.alignment = center_alignment
            cell.border = thin_border

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 15

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return {
        'avg_score': avg,
        'is_expert': is_expert,
        'answers': answers,
        'excel_file': excel_buffer
    }

async def get_available_roles():
    """Получение списка доступных ролей DAMA из базы данных"""
    try:
        async with get_async_session() as session:
            result = await session.execute(
                select(DMARoles.dama_role_name).distinct()
            )
            roles = result.scalars().all()
            return sorted(roles) if roles else []
    except Exception as e:
        logger.error(f"Ошибка при получении списка ролей: {e}")
        return []

async def get_competencies_for_role(role: str):
    """Получение компетенций для конкретной роли из базы данных"""
    try:
        if not role:
            return []

        async with get_async_session() as session:
            result = await session.execute(
                select(DAMACompetency.dama_competence_name)
                .where(DAMACompetency.dama_role_name == role)
                .distinct()
            )
            competencies = result.scalars().all()
            return sorted(competencies) if competencies else []
    except Exception as e:
        logger.error(f"Ошибка при получении компетенций для роли {role}: {e}")
        return []